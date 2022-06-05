"""
Author: Rayla Kurosaki

File: miscellaneous_functions.py

Description: This file contains all uncategorizable functions.
"""

from openpyxl import styles


def get_reverse_sorted_dict(dict):
    """
    Sorts the dictionary in reverse order according to the keys of the
    original dictionary.

    :param dict: Dictionary to sort.
    :return: The sorted dictionary in reverse order.
    """
    reverse_sorted_dict = {}
    for key in sorted(dict, reverse=True):
        reverse_sorted_dict[key] = dict[key]
    return reverse_sorted_dict


def format_num_2(num):
    """
    Format the number to 2 decimal places.

    :param num: The number to format.
    :return: The number formatted to 2 decimal places.
    """
    return float("{:.2f}".format(num))


def compute_drop_value(drop):
    """
    Computes the value of the drop. Each green material is worth 1. Each blue
    material is worth 3 green materials. Each purple material is worth 3 blue
    materials. Each gold material is worth 3 purple materials. The value of
    each drop is computed by computing its value in green materials.

    :param drop: The drop to compute the value of.
    :return: The value of the drop.
    """
    if len(drop) == 3:
        p, b, g = drop
        y = 0
    else:
        y, p, b, g = drop
    return int(g) + 3 * int(b) + 9 * int(p) + 27 * int(y)


def get_fills(data_type):
    """
    Initializes and returns list of color fills based on the type of data
    requested.

    :param data_type: Talent or Weapon material data.
    :return: A list of color fills based on the type of data requested.
    """
    gold_fill = styles.PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    purple_fill = styles.PatternFill(
        start_color="7030A0", end_color="7030A0", fill_type="solid"
    )
    blue_fill = styles.PatternFill(
        start_color="00B0F0", end_color="00B0F0", fill_type="solid"
    )
    green_fill = styles.PatternFill(
        start_color="66FF66", end_color="66FF66", fill_type="solid"
    )
    value_fill = styles.PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )
    if data_type == "Talent":
        return [purple_fill, blue_fill, green_fill, value_fill, value_fill]
    else:
        return [gold_fill, purple_fill, blue_fill, green_fill, value_fill,
                value_fill]
    pass

# def get_fills(color_dict):
#     """
#     Initializes and returns list of color fills based on the type of data
#     requested.
#
#     :param color_dict: A dictionary that contains the colors to extract from.
#     :return: A list of color fills.
#     """
#     lst = []
#     for values in color_dict.values():
#         color = values[1:]
#         lst.append(styles.PatternFill(
#             start_color=color, end_color=color, fill_type="solid"
#         ))
#         pass
#     return lst
