"""
Author: Rayla Kurosaki

File: main.py

Description: This program computes the frequency of each type of drop for the
             Talent and Weapon material drops for the game Genshin Impact.
"""

import copy
import sys
from os.path import exists as file_exists

import numpy as np
from openpyxl import formatting, styles
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

import __utils__ as utils


def get_talent_data(workbook):
    """
    Creates a dictionary to store the number of type of talent drops.

    :param workbook: The Microsoft Excel Workbook/Spreadsheet to parse
                     through.
    :return: The dictionary to store the number of type of talent drops,
             sorted in descending order.
    """
    # Get the worksheet that contains the list of talent domain drops.
    worksheet = utils.get_worksheet(workbook, "Talent Drop Data")
    # Initialize a dictionary to store the data from the worksheet.
    talent = {}
    # For each row in the spreadsheet.
    for i, row in enumerate(worksheet.iter_rows()):
        # Ignore the first row (the header row).
        if i > 0:
            # Unpack the row.
            p, b, g = row[0].value, row[1].value, row[2].value
            # Create a string top represent the drop.
            string = f"({p}/{b}/{g})"
            # If the drop is in the dictionary.
            if string in talent:
                # Increment the occurrence for that drop by 1.
                talent[string] += 1
                pass
            # If the drop is not in the dictionary.
            else:
                # Add it as a new entry to the dictionary with an occurrence
                # of one.
                talent[string] = 1
                pass
            pass
        pass
    return utils.get_reverse_sorted_dict(talent)


def get_weapon_data(workbook):
    """
    Creates a dictionary to store the number of type of weapon drops.

    :param workbook: The Microsoft Excel Workbook/Spreadsheet to parse
                     through.
    :return: The dictionary to store the number of type of weapon drops,
             sorted in descending order.
    """
    # Get the worksheet that contains the list of weapon domain drops.
    worksheet = utils.get_worksheet(workbook, "Weapon Drop Data")
    # Initialize a dictionary to store the data from the worksheet.
    weapon = {}
    # For each row in the spreadsheet.
    for i, row in enumerate(worksheet.iter_rows()):
        # Ignore the first row (the header row).
        if i > 0:
            # Unpack the row.
            y, p = row[0].value, row[1].value
            b, g = row[2].value, row[3].value
            # Create a string top represent the drop.
            string = f"({y}/{p}/{b}/{g})"
            # If the drop is in the dictionary.
            if string in weapon:
                # Increment the occurrence for that drop by 1.
                weapon[string] += 1
                pass
            # If the drop is not in the dictionary.
            else:
                # Add it as a new entry to the dictionary with an occurrence
                # of one.
                weapon[string] = 1
                pass
            pass
        pass
    return utils.get_reverse_sorted_dict(weapon)


def compute_drop_value(drop):
    """
    Computes the value of the drop. Each green material is worth 1. Each blue
    material is worth 3 green materials. Each purple material is worth 3 blue
    materials. Each gold material is worth 3 purple materials. The value of
    each drop is computed by computing its value in green materials.

    :param drop: The drop to compute the value of.
    :return: The value of the drop.
    """
    # Case where the drop is a talent material drop.
    if len(drop) == 3:
        p, b, g = drop
        # Set the gold drop to 0 since you cannot get gold talent materials.
        y = 0
    # Case where the drop is a weapon material drop.
    else:
        y, p, b, g = drop
    return int(g) + 3 * int(b) + 9 * int(p) + 27 * int(y)


def get_data_to_print(header, data):
    """
    Gets all the necessary data to print.

    :param header: First row of the table.
    :param data: Desired data to print.
    :return: The necessary data to print.
    """
    # Compute the number of runs.
    num_runs = sum(data.values())
    # Initialize the data to print.
    data_to_print = copy.deepcopy(np.array(header))
    # Iterate through each item in the dictionary.
    for k, v in data.items():
        # Compute the value of this type of drop.
        val = compute_drop_value(k[1:-1].split("/"))
        # Add t
        lst = np.array([
            [k, val, v, f"{utils.format_num_2(100 * v / num_runs)}%"]
        ])
        #
        data_to_print = np.vstack((data_to_print, lst))
    return data_to_print


def print_analysis(datas):
    """
    Prints out the analysis for the Talent material drops and for the Weapon
    material drops.

    :param datas: A list of data to print.
    """
    # Initialize the header for each section.
    sxn_headers = ["Talent Domain Drop Analysis",
                   "Weapon Domain Drop Analysis"]
    # Initialize the first row of the tables.
    table_header = [["Drop", "Value", "Occurrence", "Probability"]]
    # Write onto a file.
    with open("../data/talent_weapon_drop_analysis.txt", "w") as f:
        # For each data set.
        for i, (data_i, sub_header) in enumerate(zip(datas, sxn_headers)):
            # Get the data to print.
            data = get_data_to_print(table_header, data_i)
            # Get the maximum length of each column.
            max_len = utils.get_max_len(data)
            # Get the spacing of the section header.
            spacing = utils.get_spacing(max_len, sub_header)
            # Write the header onto the file.
            f.write(f"{'=' * spacing} {sub_header} {'=' * spacing}\n")
            # Print the boundary of the table.
            utils.print_boundary(f, max_len)
            # For each row in the data to print.
            for j, row in enumerate(data):
                # If the row is the fist non-header row.
                if j == 1:
                    # Print the line that separates the header with the
                    # analyzed data.
                    utils.print_separator(f, max_len)
                    pass
                # Print the row.
                utils.print_row(f, row, max_len)
                pass
            # Print the boundary of the table.
            utils.print_boundary(f, max_len)
            # If this is not the last set of data to print.
            if i + 1 < len(datas):
                # Print some new lines to separate the tables.
                f.write("\n\n")
            pass
        pass
    pass


def update_workbook(workbook, talent, weapon):
    worksheet_name = "Analysis"
    if utils.get_worksheet(workbook, worksheet_name):
        utils.delete_worksheet(workbook, worksheet_name)
    utils.create_new_worksheet(workbook, worksheet_name)
    worksheet = utils.get_worksheet(workbook, worksheet_name)

    col_widths = [2.77734375, 6.44140625, 4.6640625, 6.0, 5.5546875,
                  9.88671875, 2.77734375, 4.88671875, 6.44140625, 4.6640625,
                  6.0, 5.5546875, 9.88671875, 2.77734375]

    for i in range(1, 15):
        col_letter = get_column_letter(i)
        worksheet.column_dimensions[col_letter].width = col_widths[i - 1]
        pass

    black_fill = styles.PatternFill(
        start_color="000000", end_color="000000", fill_type="solid"
    )
    gold_fill = styles.PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    purple_fill = styles.PatternFill(
        start_color="7030A0", end_color="7030A0", fill_type="solid"
    )
    blue_fill = styles.PatternFill(
        start_color="00B0F0", end_color="00B0F0", fill_type="solid"
    )
    green_fill = styles.PatternFill(
        start_color="66FF66", end_color="66FF66", fill_type="solid"
    )
    value_fill = styles.PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )
    for i in range(1, 51):
        utils.get_cell(worksheet, (i, 1)).fill = black_fill
        utils.get_cell(worksheet, (i, 7)).fill = black_fill
        utils.get_cell(worksheet, (i, 14)).fill = black_fill
        if i < 15:
            utils.get_cell(worksheet, (1, i)).fill = black_fill
            pass
        pass

    font = Font(size=24)
    alignment = Alignment(horizontal="center", vertical="center")
    color_rule = ColorScaleRule(
        start_type='min', start_color="80F8696B",
        mid_type='percentile', mid_value=50, mid_color="80FFEB84",
        end_type='max', end_color="8063BE7B"
    )

    header_fill = [purple_fill, blue_fill, green_fill, value_fill,
                   value_fill]
    style_talent(worksheet, talent, font, alignment, header_fill, color_rule)

    header_fill = [gold_fill, purple_fill, blue_fill, green_fill, value_fill,
                   value_fill]
    style_weapon(worksheet, weapon, font, alignment, header_fill, color_rule)
    pass


def style_talent(worksheet, talent, font, alignment, header_fill, color_rule):
    utils.merge_cells(worksheet, "B2", "F3")
    utils.update_cell_value(worksheet, "B2", "Talent Materials")
    header_cell = utils.get_cell(worksheet, "B2")
    header_cell.font = font
    header_cell.alignment = alignment
    sub_header = ["Purple", "Blue", "Green", "Value", "Probability"]
    fonts = [Font(b=True, color="FFFFFF"), Font(b=True, color="FFFFFF"),
             Font(b=True, color="000000"), Font(size=11), Font(size=11)]
    for i, (e, font, fill) in enumerate(zip(sub_header, fonts, header_fill)):
        cell_loc = f"{get_column_letter(i + 2)}4"
        cell = utils.get_cell(worksheet, cell_loc)
        utils.update_cell_value(worksheet, cell_loc, e)
        cell.font = fonts[i]
        cell.fill = fill
        pass
    talent_data = []
    talent_runs = sum(talent.values())
    for drop, occurrence in talent.items():
        p, b, g = drop[1:-1].split("/")
        value = compute_drop_value(drop[1:-1].split("/"))
        probability = occurrence / talent_runs
        talent_data.append([int(p), int(b), int(g), value, probability])
        pass
    for i, row in enumerate(talent_data):
        for j, value in enumerate(row):
            r, c = i + 5, j + 2
            utils.update_cell_value(worksheet, (r, c), value)
            if j == len(row) - 1:
                cell = utils.get_cell(worksheet, (r, c))
                cell.number_format = '0.00%'
                pass
            pass
        pass
    start_cell, end_cell = "F5", f"F{4 + len(talent_data)}"
    utils.apply_color_scale(worksheet, start_cell, end_cell, color_rule)
    pass


def style_weapon(worksheet, weapon, font, alignment, header_fill, color_rule):
    utils.merge_cells(worksheet, "H2", "M3")
    utils.update_cell_value(worksheet, "H2", "Weapon Materials")
    header_cell = utils.get_cell(worksheet, "H2")
    header_cell.font = font
    header_cell.alignment = alignment
    sub_header = ["Gold", "Purple", "Blue", "Green", "Value", "Probability"]
    fonts = [Font(b=True, color="000000"), Font(b=True, color="FFFFFF"),
             Font(b=True, color="FFFFFF"), Font(b=True, color="000000"),
             Font(size=11), Font(size=11)]
    for i, (e, font, fill) in enumerate(zip(sub_header, fonts, header_fill)):
        cell_loc = f"{get_column_letter(i + 8)}4"
        cell = utils.get_cell(worksheet, cell_loc)
        utils.update_cell_value(worksheet, cell_loc, e)
        cell.font = fonts[i]
        cell.fill = fill
        pass
    weapon_data = []
    weapon_runs = sum(weapon.values())
    for drop, occurrence in weapon.items():
        y, p, b, g = drop[1:-1].split("/")
        value = compute_drop_value(drop[1:-1].split("/"))
        probability = occurrence / weapon_runs
        weapon_data.append([int(y), int(p), int(b), int(g), value,
                            probability])
        pass
    for i, row in enumerate(weapon_data):
        for j, value in enumerate(row):
            r, c = i + 5, j + 8
            utils.update_cell_value(worksheet, (r, c), value)
            if j == len(row) - 1:
                cell = utils.get_cell(worksheet, (r, c))
                cell.number_format = '0.00%'
                pass
            pass
        pass
    start_cell, end_cell = "M5", f"M{4 + len(weapon_data)}"
    utils.apply_color_scale(worksheet, start_cell, end_cell, color_rule)
    pass


def main():
    # Hardcode the path from source root.
    path = "../data/talent_weapon_drops.xlsx"
    # Check if the file exists
    if not file_exists(path):
        # Exits the program if the file does not exist.
        print("Place your excel file in the data directory.\n"
              "Make sure it is called \"talent_weapon_drops.xlsx\"")
        sys.exit(0)
        pass
    # Get the Excel Spreadsheet.
    workbook = utils.get_workbook(path)
    # Get the talent drops.
    talent = get_talent_data(workbook)
    # Get the weapon drops.
    weapon = get_weapon_data(workbook)
    # Update the analyzed data in the workbook.
    update_workbook(workbook, talent, weapon)
    # # Print the analysis.
    print_analysis([talent, weapon])
    # Save the Workbook.
    utils.save_workbook(workbook, "../data/talent_weapon_drops.xlsx")
    pass


if __name__ == '__main__':
    main()
    pass
