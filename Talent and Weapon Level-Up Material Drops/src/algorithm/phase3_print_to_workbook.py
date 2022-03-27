"""
Author: Rayla Kurosaki

File: phase3_print_to_workbook.py

Description: This file contains functions to analyze and print the data to a
             worksheet in a Microsoft Excel Workbook.
"""

from openpyxl import styles
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from os.path import exists as file_exists

import __utils__ as utils
import __utils__.miscellaneous_functions as misc


def get_col_width(path):
    """
    Compute the column widths of the soon-to-be-deleted output file.

    :param path: Path to the soon-to-be-deleted output file.
    :return: The column widths of the soon-to-be-deleted output file.
    """
    if not file_exists(path):
        return None
    else:
        n = 14
        col_widths = []
        workbook = utils.get_workbook(path)
        worksheet = utils.get_worksheet(workbook, "Analysis")
        for i in range(1, n + 1):
            col_letter = get_column_letter(i)
            col_widths.append(worksheet.column_dimensions[col_letter].width)
            pass
        return col_widths
    pass


def init_new_workbook(col_widths, talent, weapon):
    """
    Creates a new, styled workbook.

    :param col_widths: The width size of each column.
    :param talent: The Talent Material data to analyze.
    :param weapon: The Weapon Material data to analyze.
    :return: A new, styled workbook.
    """
    workbook = utils.create_workbook()
    worksheet_name = "Analysis"
    utils.create_new_worksheet(workbook, worksheet_name)
    worksheet = utils.get_worksheet(workbook, worksheet_name)
    if col_widths is not None:
        for i, width in enumerate(col_widths):
            col_letter = get_column_letter(i + 1)
            worksheet.column_dimensions[col_letter].width = width
            pass
        pass
    black_fill = styles.PatternFill(
        start_color="000000", end_color="000000", fill_type="solid"
    )
    for i in range(1, 100 + 1):
        utils.get_cell(worksheet, (i, 1)).fill = black_fill
        utils.get_cell(worksheet, (i, 7)).fill = black_fill
        utils.get_cell(worksheet, (i, 14)).fill = black_fill
        if i == 1:
            for j in range(1, 14 + 1):
                utils.get_cell(worksheet, (1, j)).fill = black_fill
                pass
            pass
        pass

    utils.merge_cells(worksheet, "B2", "F3")
    utils.update_cell_value(worksheet, "B2", "Talent Materials")
    header_talent_cell = utils.get_cell(worksheet, "B2")
    header_talent_cell.font = Font(size=24)
    header_talent_cell.alignment = Alignment(horizontal="center",
                                             vertical="center")
    sub_header = ["Purple", "Blue", "Green", "Value", "Probability"]
    talent_fonts = [Font(b=True, color="FFFFFF"),
                    Font(b=True, color="FFFFFF"),
                    Font(b=True, color="000000"),
                    Font(size=11), Font(size=11)]
    talent_fills = misc.get_fills("Talent")
    for i, (header, talent_font) in enumerate(zip(sub_header, talent_fonts)):
        cell_loc = f"{get_column_letter(i + 2)}4"
        utils.update_cell_value(worksheet, cell_loc, header)
        cell = utils.get_cell(worksheet, cell_loc)
        cell.font = talent_font
        pass
    for i, talent_fill in enumerate(talent_fills):
        cell_loc = f"{get_column_letter(i + 2)}4"
        cell = utils.get_cell(worksheet, cell_loc)
        cell.fill = talent_fill
        pass
    color_rule = ColorScaleRule(
        start_type='min', start_color="80F8696B",
        mid_type='percentile', mid_value=50, mid_color="80FFEB84",
        end_type='max', end_color="8063BE7B"
    )
    start_cell, end_cell = "F5", f"F{4 + len(talent)}"
    utils.apply_color_scale(worksheet, start_cell, end_cell, color_rule)

    utils.merge_cells(worksheet, "H2", "M3")
    utils.update_cell_value(worksheet, "H2", "Weapon Materials")
    header_weapon_cell = utils.get_cell(worksheet, "H2")
    header_weapon_cell.font = Font(size=24)
    header_weapon_cell.alignment = Alignment(horizontal="center",
                                             vertical="center")
    sub_header = ["Gold", "Purple", "Blue", "Green", "Value", "Probability"]
    weapon_fonts = [Font(b=True, color="000000"),
                    Font(b=True, color="FFFFFF"),
                    Font(b=True, color="FFFFFF"),
                    Font(b=True, color="000000"),
                    Font(size=11), Font(size=11)]
    weapon_fills = misc.get_fills("Weapon")
    for i, (header, weapon_font) in enumerate(zip(sub_header, weapon_fonts)):
        cell_loc = f"{get_column_letter(i + 8)}4"
        utils.update_cell_value(worksheet, cell_loc, header)
        cell = utils.get_cell(worksheet, cell_loc)
        cell.font = weapon_font
        pass
    for i, weapon_fill in enumerate(weapon_fills):
        cell_loc = f"{get_column_letter(i + 8)}4"
        cell = utils.get_cell(worksheet, cell_loc)
        cell.fill = weapon_fill
        pass
    start_cell, end_cell = "M5", f"M{4 + len(weapon)}"
    utils.apply_color_scale(worksheet, start_cell, end_cell, color_rule)
    utils.delete_worksheet(workbook, "Sheet")
    return workbook


def add_talent_data(workbook, talent):
    """
    Adds the analyzed Talent data set and prints it as a table in the Analysis
    worksheet.

    :param workbook: Current workbook to manipulate.
    :param talent: The Talent Material data to analyze.
    """
    worksheet = utils.get_worksheet(workbook, "Analysis")
    talent_data = []
    talent_runs = sum(talent.values())
    for drop, occurrence in talent.items():
        p, b, g = drop[1:-1].split("/")
        value = misc.compute_drop_value(drop[1:-1].split("/"))
        probability = occurrence / talent_runs
        talent_data.append([int(p), int(b), int(g), value, probability])
        pass
    for i, row in enumerate(talent_data):
        for j, value in enumerate(row):
            r, c = i + 5, j + 2
            utils.update_cell_value(worksheet, (r, c), value)
            if j == len(row) - 1:
                cell = utils.get_cell(worksheet, (r, c))
                cell.number_format = '0.00%'
                pass
            pass
        pass
    pass


def add_weapon_data(workbook, weapon):
    """
    Adds the analyzed Weapon data set and prints it as a table in the Analysis
    worksheet.

    :param workbook: Current workbook to manipulate.
    :param weapon: The Weapon Material data to analyze.
    """
    worksheet = utils.get_worksheet(workbook, "Analysis")
    weapon_data = []
    weapon_runs = sum(weapon.values())
    for drop, occurrence in weapon.items():
        y, p, b, g = drop[1:-1].split("/")
        value = misc.compute_drop_value(drop[1:-1].split("/"))
        probability = occurrence / weapon_runs
        weapon_data.append([int(y), int(p), int(b), int(g), value,
                            probability])
        pass
    for i, row in enumerate(weapon_data):
        for j, value in enumerate(row):
            r, c = i + 5, j + 8
            utils.update_cell_value(worksheet, (r, c), value)
            if j == len(row) - 1:
                cell = utils.get_cell(worksheet, (r, c))
                cell.number_format = '0.00%'
                pass
            pass
        pass
    pass


def phase3_main(talent, weapon):
    """
    Calls functions to print out the data in a pretty format on a Microsoft
    Excel workbook.

    :param talent: The Talent Material data to print.
    :param weapon: The Weapon Material data to print.
    """
    path = "../data/output.xlsx"
    col_widths = get_col_width(path)
    workbook = init_new_workbook(col_widths, talent, weapon)
    add_talent_data(workbook, talent)
    add_weapon_data(workbook, weapon)
    utils.save_workbook(workbook, path)
    pass
