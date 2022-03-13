"""
Author: Rayla Kurosaki

File: phase3_modify_workbook.py

Description: This file contains functions to analyze and print the data to a
             worksheet in a Microsoft Excel Workbook.
"""

from openpyxl import styles
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

import __utils__ as utils
import __utils__.miscellaneous_functions as misc


def get_analysis_worksheet(workbook, talent, weapon):
    """
    Gets the Analysis worksheet in the workbook. If the sheet deos not exist,
    then create a new Analysis, styled sheet.

    :param workbook: Current workbook to manipulate.
    :param talent: The Talent Material data to analyze.
    :param weapon: The Weapon Material data to analyze.
    :return: The Analysis worksheet.
    """
    # Initialize the name of Analysis worksheet.
    worksheet_name = "Analysis"
    # If the Analysis worksheet does not exist in the workbook.
    if not utils.worksheet_exists(workbook, worksheet_name):
        # Create a new, empty worksheet with that name.
        utils.create_new_worksheet(workbook, worksheet_name)
        worksheet = utils.get_worksheet(workbook, worksheet_name)
        # Style the newly created worksheet.
        style_analysis_worksheet(worksheet, talent, weapon)
        return worksheet
    # Otherwise, get the Analysis worksheet.
    else:
        worksheet = utils.get_worksheet(workbook, worksheet_name)
        # Remove color rules that are already present in the Analysis
        # worksheet.
        worksheet.conditional_formatting = ConditionalFormattingList()
        # Initialize a 3-color scale based on percentage.
        color_rule = ColorScaleRule(
            start_type='min', start_color="80F8696B",
            mid_type='percentile', mid_value=50, mid_color="80FFEB84",
            end_type='max', end_color="8063BE7B"
        )
        # Apply a Color scale to the potential non-empty cells in the
        # probability column fop the Talent and Weapon Material data sets.
        start_cell, end_cell = "F5", f"F{4 + len(talent)}"
        utils.apply_color_scale(worksheet, start_cell, end_cell, color_rule)
        start_cell, end_cell = "M5", f"M{4 + len(weapon)}"
        utils.apply_color_scale(worksheet, start_cell, end_cell, color_rule)
        return worksheet
    pass


def get_fills(data_type):
    """
    Initializes and returns list of color fills based on the type of data
    requested.

    :param data_type: Talent or Weapon material data.
    :return: A list of color fills based on the type of data requested.
    """
    # Initialize some color fills.
    gold_fill = styles.PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    purple_fill = styles.PatternFill(
        start_color="7030A0", end_color="7030A0", fill_type="solid"
    )
    blue_fill = styles.PatternFill(
        start_color="00B0F0", end_color="00B0F0", fill_type="solid"
    )
    green_fill = styles.PatternFill(
        start_color="66FF66", end_color="66FF66", fill_type="solid"
    )
    value_fill = styles.PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )
    # Return a list of color fills based on the type of data requested.
    if data_type == "Talent":
        return [purple_fill, blue_fill, green_fill, value_fill, value_fill]
    else:
        return [gold_fill, purple_fill, blue_fill, green_fill, value_fill,
                value_fill]


def style_analysis_worksheet(worksheet, talent, weapon):
    """
    Styles the Analysis worksheet.

    :param worksheet: Current worksheet to manipulate.
    :param talent: The Talent Material data to analyze.
    :param weapon: The Weapon Material data to analyze.
    """
    # Set the tab color to black.
    worksheet.sheet_properties.tabColor = "000000"
    # Initialize column width sizes.
    col_widths = [2.77734375, 6.44140625, 4.6640625, 6.0, 5.5546875,
                  9.88671875, 2.77734375, 4.88671875, 6.44140625, 4.6640625,
                  6.0, 5.5546875, 9.88671875, 2.77734375]
    # Set the column sizes of the Analysis worksheet.
    for i, width in enumerate(col_widths):
        col_letter = get_column_letter(i + 1)
        worksheet.column_dimensions[col_letter].width = width
        pass
    # Initialize a black fill color.
    black_fill = styles.PatternFill(
        start_color="000000", end_color="000000", fill_type="solid"
    )
    # For a set number of rows.
    for i in range(1, 100 + 1):
        # Fill the cell in columns A, G, and N with black.
        utils.get_cell(worksheet, (i, 1)).fill = black_fill
        utils.get_cell(worksheet, (i, 7)).fill = black_fill
        utils.get_cell(worksheet, (i, 14)).fill = black_fill
        # If this is the first row.
        if i == 1:
            # For the first 14 cells.
            for j in range(1, 14 + 1):
                # Fill the cell with black.
                utils.get_cell(worksheet, (1, j)).fill = black_fill
                pass
            pass
        pass

    # Merge some cells for the Talent Material Main header.
    utils.merge_cells(worksheet, "B2", "F3")
    # Set the main header title for the Talent Material data.
    utils.update_cell_value(worksheet, "B2", "Talent Materials")
    # Get the Talent Header Cell
    header_talent_cell = utils.get_cell(worksheet, "B2")
    # Set the font size to 24
    header_talent_cell.font = Font(size=24)
    # Center this text in this cell horizontally and vertically.
    header_talent_cell.alignment = Alignment(horizontal="center",
                                             vertical="center")
    # Initialize the sub header for the Talent material data.
    sub_header = ["Purple", "Blue", "Green", "Value", "Probability"]
    # Initialize the Fonts of the sub headers for the Talent Material data.
    talent_fonts = [Font(b=True, color="FFFFFF"),
                    Font(b=True, color="FFFFFF"),
                    Font(b=True, color="000000"),
                    Font(size=11), Font(size=11)]
    # Initialize the fills of the sub headers for the Talent Material data.
    talent_fills = get_fills("Talent")
    # Set the 4-th row as the sub header for the Talent Material data.
    # Apply font styling to the sub header for the Talent Material data.
    for i, (header, talent_font) in enumerate(zip(sub_header, talent_fonts)):
        cell_loc = f"{get_column_letter(i + 2)}4"
        utils.update_cell_value(worksheet, cell_loc, header)
        cell = utils.get_cell(worksheet, cell_loc)
        cell.font = talent_font
        pass
    # Apply fill styling to the sub header for the Talent Material data.
    for i, talent_fill in enumerate(talent_fills):
        cell_loc = f"{get_column_letter(i + 2)}4"
        cell = utils.get_cell(worksheet, cell_loc)
        cell.fill = talent_fill
        pass
    # Initialize a 3-color scale based on percentage.
    color_rule = ColorScaleRule(
        start_type='min', start_color="80F8696B",
        mid_type='percentile', mid_value=50, mid_color="80FFEB84",
        end_type='max', end_color="8063BE7B"
    )
    # Apply a Color scale to the potential non-empty cells in the probability
    # column.
    start_cell, end_cell = "F5", f"F{4 + len(talent)}"
    utils.apply_color_scale(worksheet, start_cell, end_cell, color_rule)

    # Merge some cells for the Weapon Material Main header.
    utils.merge_cells(worksheet, "H2", "M3")
    # Set the main header title for the Weapon Material data.
    utils.update_cell_value(worksheet, "H2", "Weapon Materials")
    # Get the Weapon Header Cell
    header_weapon_cell = utils.get_cell(worksheet, "H2")
    # Set the font size to 24
    header_weapon_cell.font = Font(size=24)
    # Center this text in this cell horizontally and vertically.
    header_weapon_cell.alignment = Alignment(horizontal="center",
                                             vertical="center")
    # Initialize the sub header for the Weapon material data.
    sub_header = ["Gold", "Purple", "Blue", "Green", "Value", "Probability"]
    # Initialize the Fonts of the sub headers for the Weapon Material data.
    weapon_fonts = [Font(b=True, color="000000"),
                    Font(b=True, color="FFFFFF"),
                    Font(b=True, color="FFFFFF"),
                    Font(b=True, color="000000"),
                    Font(size=11), Font(size=11)]
    # Initialize the fills of the sub headers for the Weapon Material data.
    weapon_fills = get_fills("Weapon")
    # Set the 4-th row as the sub header for the Weapon Material data.
    for i, (header, weapon_font) in enumerate(zip(sub_header, weapon_fonts)):
        cell_loc = f"{get_column_letter(i + 8)}4"
        utils.update_cell_value(worksheet, cell_loc, header)
        cell = utils.get_cell(worksheet, cell_loc)
        cell.font = weapon_font
        pass
    # Apply fill styling to the sub header for the Weapon Material data.
    for i, weapon_fill in enumerate(weapon_fills):
        cell_loc = f"{get_column_letter(i + 8)}4"
        cell = utils.get_cell(worksheet, cell_loc)
        cell.fill = weapon_fill
        pass
    # Apply a Color scale to the potential non-empty cells in the probability
    # column.
    start_cell, end_cell = "M5", f"M{4 + len(weapon)}"
    utils.apply_color_scale(worksheet, start_cell, end_cell, color_rule)
    pass


def add_talent_data(worksheet, talent):
    """
    Adds the analyzed Talent data set and prints it as a table in the Analysis
    worksheet.

    :param worksheet: Current worksheet to manipulate.
    :param talent: The Talent Material data to analyze.
    """
    # Initializes the set of data to print.
    talent_data = []
    # Get the number of Talent domain.
    talent_runs = sum(talent.values())
    # For each unique talent domain drop.
    for drop, occurrence in talent.items():
        # Unpack the drop.
        p, b, g = drop[1:-1].split("/")
        # Compute the value of the drop.
        value = misc.compute_drop_value(drop[1:-1].split("/"))
        # Compute the probability of that type of drop occurring.
        probability = occurrence / talent_runs
        # Create a list of data to print from this drop and append it to the
        # main list.
        talent_data.append([int(p), int(b), int(g), value, probability])
        pass
    # For each drop.
    for i, row in enumerate(talent_data):
        # For each element in the row.
        for j, value in enumerate(row):
            # Update the cell with that value.
            r, c = i + 5, j + 2
            utils.update_cell_value(worksheet, (r, c), value)
            # For the last element of the row.
            if j == len(row) - 1:
                # Format the cell as a percentage with 2 decimal places.
                cell = utils.get_cell(worksheet, (r, c))
                cell.number_format = '0.00%'
                pass
            pass
        pass
    pass


def add_weapon_data(worksheet, weapon):
    """
    Adds the analyzed Weapon data set and prints it as a table in the Analysis
    worksheet.

    :param worksheet: Current worksheet to manipulate.
    :param weapon: The Weapon Material data to analyze.
    """
    # Initializes the set of data to print.
    weapon_data = []
    # Get the number of Weapon domain runs.
    weapon_runs = sum(weapon.values())
    # For each unique weapon domain drop.
    for drop, occurrence in weapon.items():
        # Unpack the drop.
        y, p, b, g = drop[1:-1].split("/")
        # Compute the value of the drop.
        value = misc.compute_drop_value(drop[1:-1].split("/"))
        # Compute the probability of that type of drop occurring.
        probability = occurrence / weapon_runs
        # Create a list of data to print from this drop and append it to the
        # main list.
        weapon_data.append([int(y), int(p), int(b), int(g), value,
                            probability])
        pass
    # For each drop.
    for i, row in enumerate(weapon_data):
        # For each element in the row.
        for j, value in enumerate(row):
            # Update the cell with that value.
            r, c = i + 5, j + 8
            utils.update_cell_value(worksheet, (r, c), value)
            # For the last element of the row.
            if j == len(row) - 1:
                # Format the cell as a percentage with 2 decimal places.
                cell = utils.get_cell(worksheet, (r, c))
                cell.number_format = '0.00%'
                pass
            pass
        pass
    pass


def phase3_main(workbook, talent, weapon):
    """
    Calls functions to analyze and modify the workbook.

    :param workbook: Current workbook to manipulate.
    :param talent: The Talent Material data to analyze.
    :param weapon: The Weapon Material data to analyze.
    """
    # If an Analysis sheet doesn't exist in the workbook, create a new one
    # and style it.
    analysis = get_analysis_worksheet(workbook, talent, weapon)
    # Add the Talent data
    add_talent_data(analysis, talent)
    # Add the Weapon data
    add_weapon_data(analysis, weapon)
    pass
